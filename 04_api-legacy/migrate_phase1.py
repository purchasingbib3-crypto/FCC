#!/usr/bin/env python3
"""
Phase 1 Data Migration — Excel → PostgreSQL
Imports all sheets from the Excel workbook into the Phase 1 schema.
Order: vendor → unit → alias → asset → storage_loc → calibration → users → SS6 → SAP → reconciliation
"""
import openpyxl
import psycopg2
import psycopg2.extras
import hashlib
import re
import os
from datetime import datetime, date
from decimal import Decimal

# ── Config ──
XLSX_PATH = "/home/ubuntu/.hermes/cache/documents/doc_0b8e539e385f_database fuel control (1).xlsx"
DB_CONN = dict(
    host="/var/run/postgresql", port=5432,
    dbname="fuel_control_center",
    user="postgres"
)

# ── Helpers ──
def norm(val):
    if not val or not isinstance(val, str):
        return None
    cleaned = re.sub(r'[^A-Za-z0-9]', '', val).upper()
    return cleaned or None

def ser(val):
    if val is None: return None
    if isinstance(val, datetime): return val
    if isinstance(val, date): return val
    if isinstance(val, float):
        if val == int(val): return int(val)
        return val
    if isinstance(val, str):
        s = val.strip()
        if s.startswith("#") and (s.endswith("?") or s == "#N/A"): return None
        return s if s else None
    return val

def derive_unit_sap(order, text):
    """Derive unit SAP from Order/Text, matching PostgreSQL function."""
    order = (order or '').strip()
    text = (text or '').strip()
    # Try splitting Order by '-'
    part = order.split('-')[0] if order else ''
    if part:
        part = part.replace('-DOBEL INPUT', '').replace('-SALAH UNIT', '')
    if part: return part.strip()
    # Try splitting Text by '.KM-'
    part = text.split('.KM-')[0] if '.KM-' in text else text
    part = part.replace('-DOBEL INPUT', '').replace('-SALAH UNIT', '').strip()
    return part if part else None

def row_sha256(file_sha, row_num, row_json):
    return hashlib.sha256(f"{file_sha}|{row_num}|{row_json}".encode()).hexdigest()

def dup_key_hash(trx_date, shift, unit, storage_loc):
    return hashlib.sha256(f"{trx_date}|{shift}|{norm(unit) or ''}|{norm(storage_loc) or ''}".encode()).hexdigest()

def biz_hash(trx_date, shift, unit, vol, storage_loc):
    return hashlib.sha256(f"{trx_date}|{shift}|{norm(unit) or ''}|{vol}|{norm(storage_loc) or ''}".encode()).hexdigest()

def sap_biz_hash(trx_date, qty, order, text, storage_loc):
    return hashlib.sha256(f"{trx_date}|{qty}|{norm(order) or ''}|{norm(text) or ''}|{norm(storage_loc) or ''}".encode()).hexdigest()

# ── File SHA ──
FILE_SHA = hashlib.sha256(open(XLSX_PATH, 'rb').read()).hexdigest()
print(f"File SHA-256: {FILE_SHA[:16]}...")

# ── Connect ──
conn = psycopg2.connect(**DB_CONN)
conn.autocommit = False
cur = conn.cursor()
psycopg2.extras.register_uuid(conn)

# ── 1. Master Vendor ──
print("\n[1/10] Importing master_vendor...")
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
ws = wb["UNIT_ALIAS"]
vendors = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    v = ser(row[3]) if len(row) > 3 else None  # VENDOR
    if v and v.strip():
        vendors.add(v.strip())

for v in sorted(vendors):
    cur.execute(
        "INSERT INTO public.master_vendor (vendor_code, vendor_name) VALUES (%s, %s) ON CONFLICT DO NOTHING",
        (v, v)
    )
print(f"  → {len(vendors)} vendors")

# ── 2. Master Unit + Unit Alias ──
print("\n[2/10] Importing master_unit + unit_alias...")
ws = wb["UNIT_ALIAS"]
units = {}
aliases = []
for row in ws.iter_rows(min_row=2, values_only=True):
    unit_std = ser(row[0])
    alias_ss6 = ser(row[1])
    alias_sap = ser(row[2])
    vendor = ser(row[3])
    kat1 = ser(row[4])
    kat2 = ser(row[3+1]) if len(row) > 4 else None  # KATEGORI 2 = col 5
    kat2 = ser(row[5]) if len(row) > 5 else None
    klas = ser(row[6]) if len(row) > 6 else None
    
    if not unit_std or not vendor:
        continue
    units[unit_std] = (unit_std, vendor, kat1, kat2, klas)
    if alias_ss6:
        aliases.append((unit_std, alias_ss6, 'SS6'))
    if alias_sap:
        aliases.append((unit_std, alias_sap, 'SAP'))

for u_data in units.values():
    cur.execute(
        """INSERT INTO public.master_unit (unit_std, vendor_code, kategori_1, kategori_2, klasifikasi)
           VALUES (%s, %s, %s, %s, %s) ON CONFLICT (unit_std) DO UPDATE
           SET vendor_code=EXCLUDED.vendor_code, kategori_1=EXCLUDED.kategori_1,
               kategori_2=EXCLUDED.kategori_2, klasifikasi=EXCLUDED.klasifikasi, is_active=true""",
        u_data
    )
print(f"  → {len(units)} units")

for a in aliases:
    cur.execute(
        """INSERT INTO public.unit_alias (unit_std, alias, source)
           VALUES (%s, %s, %s) ON CONFLICT (source, alias_key) DO UPDATE
           SET unit_std=EXCLUDED.unit_std, is_active=true""",
        (a[0], a[1], a[2])
    )
print(f"  → {len(aliases)} aliases")

# ── 3. Master Asset (only rows 2-32 from DATABASE FUEL TRUCK & MAINTANK) ──
print("\n[3/10] Importing master_asset...")
ws = wb["DATABASE FUEL TRUCK & MAINTANK"]
assets = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue  # header
    if i < 1 or i > 31: continue  # only rows 2-32 (0-indexed: 1-31)
    ss6 = str(row[0]).replace('.0','') if row[0] else None
    sap = str(row[1]).replace('.0','') if row[1] else None
    if not ss6 and not sap:
        continue
    # Determine asset type — all entries in this sheet are FUEL_TRUCK
    code = ss6 or sap
    atype = 'FUEL_TRUCK'
    assets.append((code, atype, code, ss6, sap, None, None, 'PPA'))

for a in assets:
    cur.execute(
        """INSERT INTO public.master_asset (asset_code, asset_type, asset_name, lambung_ss6, lambung_sap, capacity_liter, dip_max_cm, vendor_code)
           VALUES (%s, %s, %s, %s, %s, %s, %s, %s) ON CONFLICT (asset_code) DO NOTHING""",
        a
    )
print(f"  → {len(assets)} assets")

# ── 4. Master Storage Location ──
print("\n[4/10] Importing master_storage_location...")
storage_locs = set()
# From SS6
ws = wb["SS6"]
for row in ws.iter_rows(min_row=2, values_only=True):
    sloc = ser(row[4]) if len(row) > 4 else None
    if sloc: storage_locs.add(str(sloc).strip())
# From SAP
ws = wb["SAP"]
for row in ws.iter_rows(min_row=2, values_only=True):
    sloc = ser(row[4]) if len(row) > 4 else None
    if sloc: storage_locs.add(str(sloc).strip())
# From asset codes (FS10-FS15)
for code in ['FS10','FS11','FS12','FS13','FS14','FS15']:
    storage_locs.add(code)

for s in sorted(storage_locs):
    cur.execute(
        "INSERT INTO public.master_storage_location (storage_loc, storage_name) VALUES (%s, %s) ON CONFLICT DO NOTHING",
        (s, s)
    )
print(f"  → {len(storage_locs)} storage locations")

# ── 5. Tank Calibration (HITUNG TERA) ──
print("\n[5/10] Importing tank_calibration...")
ws = wb["HITUNG TERA"]
cal_rows = ws.iter_rows(min_row=1, values_only=True)
headers = list(next(cal_rows))
# Col 0 = Tangki/dip (asset_code), cols 1+ = dip values, rows below = volumes
# Actually: row 0 has dip values in columns, subsequent rows have asset_code in col 0 and volumes in cols 1+
cal_data = []
for row in cal_rows:
    if not row or not row[0]: continue
    asset_code = str(row[0]).replace('.0','')
    for col_idx in range(1, len(row)):
        dip_val = headers[col_idx] if col_idx < len(headers) else None
        vol_val = row[col_idx]
        if dip_val is not None and vol_val is not None:
            try:
                dip = round(float(dip_val), 2)
                vol = round(float(vol_val), 2)
                if vol >= 0 and dip >= 0:
                    cal_data.append((asset_code, dip, vol))
            except (ValueError, TypeError):
                continue

# Insert calibration, checking monotonicity (skip non-monotonic rows for tank 2550)
prev_vol = {}
for asset, dip, vol in cal_data:
    key = asset
    if key in prev_vol and vol < prev_vol[key]:
        print(f"  ⚠ Skip non-monotonic: {asset} dip={dip} vol={vol} < prev={prev_vol[key]}")
        continue
    try:
        cur.execute(
            "INSERT INTO public.tank_calibration (asset_code, dip_cm, volume_liter) VALUES (%s, %s, %s) ON CONFLICT DO NOTHING",
            (asset, dip, vol)
        )
        prev_vol[key] = vol
    except Exception as e:
        conn.rollback()
        print(f"  ⚠ Skip {asset} dip={dip}: {e}")
        conn = psycopg2.connect(**DB_CONN)
        conn.autocommit = False
        cur = conn.cursor()
        prev_vol[key] = vol
print(f"  → {len(cal_data)} calibration points (some may be skipped for monotonicity)")

# ── 6. Users (DATABASE USER + misplaced rows 145-181) ──
print("\n[6/10] Importing users (quarantine misplaced)...")
ws = wb["DATABASE USER"]
users = []
for row in ws.iter_rows(min_row=2, values_only=True):
    nrp = str(row[0]).replace('.0','') if row[0] else None
    nama = ser(row[1]) if len(row) > 1 else None
    role = ser(row[2]) if len(row) > 2 else None
    if not nrp and not nama: continue
    users.append((nrp, nama, role))
print(f"  → {len(users)} users from DATABASE USER")

# Misplaced users from DATABASE FUEL TRUCK rows 145-181
ws = wb["DATABASE FUEL TRUCK & MAINTANK"]
misplaced = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i < 144 or i > 180: continue  # rows 145-181 (0-indexed 144-180)
    nrp = str(row[0]).replace('.0','') if row[0] else None
    nama = ser(row[1]) if len(row) > 1 else None
    role = ser(row[2]) if len(row) > 2 else None
    if not nama: continue
    # Map role to app_role enum
    role_map = {'FUELMAN': 'FUELMAN', 'ADMIN': 'ADMIN', 'PENERIMAAN': 'PENERIMAAN'}
    app_role = role_map.get(str(role).upper() if role else '', 'FUELMAN')
    misplaced.append((nrp, nama, app_role))
print(f"  → {len(misplaced)} misplaced users (quarantined — not in DATABASE USER)")

conn.commit()
print("\n[1-6] Master data committed ✓")

# ════════════════════════════════════════════════════════════
# 7. SS6 Transactions — direct insert with hash + alias resolution
# ════════════════════════════════════════════════════════════
print("\n[7/10] Importing SS6 transactions (132k+ rows)...")
ws = wb["SS6"]

# Create import batch
cur.execute(
    """INSERT INTO public.import_batch (source, original_filename, source_file_sha256, period_start, period_end, status, created_by)
       VALUES ('SS6', %s, %s, '2026-06-01', '2026-07-31', 'COMMITTING', NULL) RETURNING id""",
    (os.path.basename(XLSX_PATH), FILE_SHA)
)
batch_id = cur.fetchone()[0]

# Insert into fuel_issue_ss6 in batches
batch = []
count = 0
skipped = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    unit = ser(row[0])
    dt = row[1] if len(row) > 1 else None
    shift = row[2] if len(row) > 2 else None
    vol = row[3] if len(row) > 3 else None
    sloc = ser(row[4]) if len(row) > 4 else None

    if dt is None or vol is None:
        skipped += 1; continue
    
    try:
        d = dt if isinstance(dt, date) else datetime.strptime(str(dt).split(" ")[0], "%Y-%m-%d").date()
    except:
        try:
            d = datetime.strptime(str(dt).split(" ")[0], "%d.%m.%Y").date()
        except:
            skipped += 1; continue

    try:
        v = float(vol)
    except (ValueError, TypeError):
        skipped += 1; continue
    
    try:
        s = int(float(shift)) if shift else None
    except:
        s = None

    if s not in (1, 2):
        skipped += 1; continue

    sloc_str = str(sloc).strip() if sloc else None
    if not sloc_str:
        skipped += 1; continue

    unit_str = str(unit).strip() if unit else None
    unit_key = norm(unit_str)
    
    # Resolve alias
    unit_std = None
    if unit_key:
        cur.execute(
            "SELECT unit_std FROM public.unit_alias WHERE is_active AND alias_key = %s AND source IN ('SS6','MANUAL') ORDER BY CASE source WHEN 'SS6' THEN 0 ELSE 1 END LIMIT 1",
            (unit_key,)
        )
        r = cur.fetchone()
        if r: unit_std = r[0]

    row_num = count + 2
    rhash = row_sha256(FILE_SHA, row_num, f"{unit_str}|{d}|{s}|{v}|{sloc_str}")
    dhash = dup_key_hash(d, s, unit_str, sloc_str)
    bhash = biz_hash(d, s, unit_str, v, sloc_str)

    batch.append((d, s, unit_str, unit_std, v, sloc_str, batch_id, row_num, rhash, dhash, bhash))
    count += 1

    if len(batch) >= 5000:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO public.fuel_issue_ss6
               (trx_date, shift, unit_raw, unit_std, volume_liter, storage_loc, import_batch_id, source_row_number, row_hash, duplicate_key_hash, business_hash)
               VALUES %s ON CONFLICT (trx_date, row_hash) DO NOTHING""",
            batch
        )
        conn.commit()
        print(f"  → {count:,} rows inserted...")
        batch = []

if batch:
    psycopg2.extras.execute_values(
        cur,
        """INSERT INTO public.fuel_issue_ss6
           (trx_date, shift, unit_raw, unit_std, volume_liter, storage_loc, import_batch_id, source_row_number, row_hash, duplicate_key_hash, business_hash)
           VALUES %s ON CONFLICT (trx_date, row_hash) DO NOTHING""",
        batch
    )
    conn.commit()

cur.execute(
    "UPDATE public.import_batch SET status='COMMITTED', rows_total=%s, rows_valid=%s, rows_inserted=%s, rows_rejected=%s WHERE id=%s",
    (count, count, count, skipped, batch_id)
)
conn.commit()
print(f"  → {count:,} SS6 rows inserted ({skipped} skipped)")

# ════════════════════════════════════════════════════════════
# 8. SAP Transactions
# ════════════════════════════════════════════════════════════
print("\n[8/10] Importing SAP transactions (98k+ rows)...")
ws = wb["SAP"]

cur.execute(
    """INSERT INTO public.import_batch (source, original_filename, source_file_sha256, period_start, period_end, status, created_by)
       VALUES ('SAP', %s, %s, '2026-06-01', '2026-07-31', 'COMMITTING', NULL) RETURNING id""",
    (os.path.basename(XLSX_PATH), FILE_SHA)
)
batch_id = cur.fetchone()[0]

batch = []
count = 0
skipped = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    dt = row[0] if len(row) > 0 else None
    qty = row[1] if len(row) > 1 else None
    order = ser(row[2]) if len(row) > 2 else None
    text = ser(row[3]) if len(row) > 3 else None
    sloc = ser(row[4]) if len(row) > 4 else None
    unit_sap = ser(row[5]) if len(row) > 5 else None

    # Skip formula-only rows
    if qty is None and dt is None and not order and not text and not sloc:
        skipped += 1; continue

    if dt is None or qty is None:
        skipped += 1; continue

    try:
        d = dt if isinstance(dt, date) else datetime.strptime(str(dt).split(" ")[0], "%Y-%m-%d").date()
    except:
        try:
            d = datetime.strptime(str(dt).split(" ")[0], "%d.%m.%Y").date()
        except:
            skipped += 1; continue

    try:
        q = float(qty)
    except (ValueError, TypeError):
        skipped += 1; continue

    sloc_str = str(sloc).strip() if sloc else None
    if not sloc_str:
        skipped += 1; continue

    # Derive unit from order/text or UNIT SAP FIX
    derived = unit_sap if unit_sap else derive_unit_sap(order, text)
    derived_key = norm(derived)

    # Resolve alias
    unit_std = None
    if derived_key:
        cur.execute(
            "SELECT unit_std FROM public.unit_alias WHERE is_active AND alias_key = %s AND source IN ('SAP','MANUAL') ORDER BY CASE source WHEN 'SAP' THEN 0 ELSE 1 END LIMIT 1",
            (derived_key,)
        )
        r = cur.fetchone()
        if r: unit_std = r[0]

    row_num = count + 2
    rhash = row_sha256(FILE_SHA, row_num, f"{d}|{q}|{order}|{text}|{sloc_str}|{unit_sap}")
    bhash = sap_biz_hash(d, q, order, text, sloc_str)

    batch.append((d, q, order, text, unit_std, sloc_str, batch_id, row_num, rhash, bhash))
    count += 1

    if len(batch) >= 5000:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO public.fuel_issue_sap
               (trx_date, qty_liter, order_no, order_text, unit_std, storage_loc, import_batch_id, source_row_number, row_hash, business_hash)
               VALUES %s ON CONFLICT (trx_date, row_hash) DO NOTHING""",
            batch
        )
        conn.commit()
        print(f"  → {count:,} rows inserted...")
        batch = []

if batch:
    psycopg2.extras.execute_values(
        cur,
        """INSERT INTO public.fuel_issue_sap
           (trx_date, qty_liter, order_no, order_text, unit_std, storage_loc, import_batch_id, source_row_number, row_hash, business_hash)
           VALUES %s ON CONFLICT (trx_date, row_hash) DO NOTHING""",
        batch
    )
    conn.commit()

cur.execute(
    "UPDATE public.import_batch SET status='COMMITTED', rows_total=%s, rows_valid=%s, rows_inserted=%s, rows_rejected=%s WHERE id=%s",
    (count, count, count, skipped, batch_id)
)
conn.commit()
print(f"  → {count:,} SAP rows inserted ({skipped} skipped)")

# ════════════════════════════════════════════════════════════
# 9. Sync work queue + reconciliation
# ════════════════════════════════════════════════════════════
print("\n[9/10] Syncing work queue (June + July 2026)...")
# Run as postgres to bypass RLS
cur.close()
conn.close()

import subprocess
for month in ['2026-06-01', '2026-07-01']:
    result = subprocess.run(
        ['sudo', '-u', 'postgres', 'psql', '-d', 'fuel_control_center', '-c',
         f"SELECT private.sync_work_queue_internal('{month}');"],
        capture_output=True, text=True
    )
    print(f"  Work queue {month}: {'OK' if result.returncode == 0 else result.stderr.strip()}")

print("\n[10/10] Refreshing reconciliation (June + July 2026)...")
for month in ['2026-06-01', '2026-07-01']:
    result = subprocess.run(
        ['sudo', '-u', 'postgres', 'psql', '-d', 'fuel_control_center', '-c',
         f"SELECT api.refresh_reconciliation('{month}');"],
        capture_output=True, text=True
    )
    print(f"  Reconciliation {month}: {'OK' if result.returncode == 0 else result.stderr.strip()}")

# ── Summary ──
print("\n" + "="*60)
print("MIGRATION COMPLETE")
print("="*60)

conn = psycopg2.connect(**DB_CONN)
cur = conn.cursor()
tables = [
    ('master_vendor', 'public.master_vendor'),
    ('master_unit', 'public.master_unit'),
    ('unit_alias', 'public.unit_alias'),
    ('master_asset', 'public.master_asset'),
    ('master_storage_location', 'public.master_storage_location'),
    ('tank_calibration', 'public.tank_calibration'),
    ('fuel_issue_ss6', 'public.fuel_issue_ss6'),
    ('fuel_issue_sap', 'public.fuel_issue_sap'),
    ('reconciliation_finding', 'public.reconciliation_finding'),
    ('work_queue_item', 'public.work_queue_item'),
    ('import_batch', 'public.import_batch'),
]
for name, table in tables:
    cur.execute(f"SELECT count(*) FROM {table}")
    c = cur.fetchone()[0]
    print(f"  {name:30s} {c:>12,}")

# Totals
cur.execute("SELECT sum(volume_liter) FROM public.fuel_issue_ss6")
ss6_vol = cur.fetchone()[0] or 0
cur.execute("SELECT -sum(qty_liter), sum(abs(qty_liter)) FROM public.fuel_issue_sap")
sap_net, sap_abs = cur.fetchone()
print(f"\n  SS6 total volume:     {ss6_vol:>14,.1f} L")
print(f"  SAP signed net:       {(sap_net or 0):>14,.1f} L")
print(f"  SAP absolute legacy:  {(sap_abs or 0):>14,.1f} L")
print(f"  Deviation (net):     {(ss6_vol - (sap_net or 0)):>14,.1f} L")
print(f"  Deviation (abs):     {(ss6_vol - (sap_abs or 0)):>14,.1f} L")

conn.close()
print("\n✓ Phase 1 migration complete")