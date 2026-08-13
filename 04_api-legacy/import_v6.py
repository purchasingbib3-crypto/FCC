#!/usr/bin/env python3
"""
Import Excel V6 ke schema fcc PostgreSQL.
"""
import openpyxl
import psycopg2
import psycopg2.extras
import re
from datetime import datetime, date
from pathlib import Path

XLSX = "/home/ubuntu/.hermes/cache/documents/doc_15de56870ef7_Fuel_Control_Center_Database_Template_V6(1).xlsx"
DSN = dict(host="/var/run/postgresql", port=5432, dbname="fuel_control_center", user="postgres")

def ser(v):
    if v is None: return None
    if isinstance(v, datetime): return v
    if isinstance(v, date): return v
    if isinstance(v, float):
        if v == int(v): return int(v)
        return v
    if isinstance(v, str):
        s = v.strip()
        if s.startswith("#") and (s.endswith("?") or s == "#N/A"): return None
        return s if s else None
    return v

def fix_vendor(v):
    if v in (None, "", "Vendor", "SARANA BOSTON"):
        return "PPA"
    return v

def get_rows(ws, header_row=4):
    """Skip header rows, return dict per row."""
    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < header_row - 1: continue
        if headers is None:
            headers = [ser(c) for c in row]
            continue
        if all(c is None or c == '' for c in row): continue
        rec = {}
        for h, v in zip(headers, row):
            if h: rec[h] = ser(v)
        rows.append(rec)
    return rows

# Connect
conn = psycopg2.connect(**DSN)
conn.autocommit = False
cur = conn.cursor()

wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)

# 1. Users
print("[1/14] Importing users...")
ws = wb['03_USERS']
rows = get_rows(ws, header_row=4)
import hashlib
for r in rows:
    if not r.get('NRP / Username'): continue
    nrp = str(r.get('NRP / Username'))
    pwd = str(r.get('Password Awal') or 'demo123')
    pwd_hash = hashlib.sha256(pwd.encode()).hexdigest()
    cur.execute("""
        INSERT INTO fcc.app_user (username, nama, role, vendor_kode, status, password_hash, must_change_pw)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (upper(username)) DO UPDATE SET nama=EXCLUDED.nama, role=EXCLUDED.role
    """, (nrp, r.get('Nama'), r.get('Role'), fix_vendor(r.get('Vendor')),
          r.get('Status') or 'ACTIVE', pwd_hash,
          str(r.get('Wajib Ganti Password', 'YA')).upper() == 'YA'))
conn.commit()
print(f"  → {cur.rowcount} users")

# 2. Master Vendor
print("\n[2/14] Importing master_vendor...")
ws = wb['05_MASTER_VENDOR']
rows = get_rows(ws)
for r in rows:
    if not r.get('Kode Vendor'): continue
    cur.execute("""
        INSERT INTO fcc.master_vendor (kode, nama, kategori, status)
        VALUES (%s, %s, %s, %s) ON CONFLICT (kode) DO UPDATE SET nama=EXCLUDED.nama
    """, (r.get('Kode Vendor'), r.get('Nama Vendor'), r.get('Kategori', 'INTERNAL'),
          r.get('Status', 'ACTIVE')))
conn.commit()
print(f"  → {cur.rowcount} vendors")

# 3. Master Unit
print("\n[3/14] Importing master_unit...")
ws = wb['04_MASTER_UNIT']
rows = get_rows(ws)
imported_units = set()
for r in rows:
    kode = r.get('Kode Unit')
    if not kode or kode in ('Kode Unit', 'Unit Standar'): continue
    if kode in imported_units: continue
    imported_units.add(kode)
    cur.execute("""
        INSERT INTO fcc.master_unit (kode, nama, vendor_kode, kategori, status)
        VALUES (%s, %s, %s, %s, %s) ON CONFLICT (kode) DO UPDATE SET nama=EXCLUDED.nama
    """, (r.get('Kode Unit'), r.get('Nama Unit') or r.get('Kode Unit'), fix_vendor(r.get('Vendor')),
          r.get('Kategori', 'SARANA'), r.get('Status', 'ACTIVE')))
conn.commit()
print(f"  → {cur.rowcount} units")

# 4. Unit Alias
print("\n[4/14] Importing unit_alias...")
ws = wb['11_UNIT_ALIAS']
rows = get_rows(ws)
imported_aliases = set()
for r in rows:
    us = r.get('Unit Standar')
    if not us or us in ('Unit Standar', 'Kode Unit'): continue
    if us not in imported_units: continue  # FK check
    key = (us, r.get('Alias SS6'), r.get('Alias SAP'))
    if key in imported_aliases: continue
    imported_aliases.add(key)
    vendor_a = fix_vendor(r.get('Vendor'))
    if vendor_a in ('Vendor', 'SARANA BOSTON') or not vendor_a:
        vendor_a = 'PPA'
    cur.execute("""
        INSERT INTO fcc.unit_alias (unit_standar, alias_ss6, alias_sap, vendor_kode, kategori, status)
        VALUES (%s, %s, %s, %s, %s, %s)
        ON CONFLICT DO NOTHING
    """, (us, r.get('Alias SS6'), r.get('Alias SAP'),
          vendor_a, r.get('Kategori'), r.get('Status', 'ACTIVE')))
conn.commit()
print(f"  → {cur.rowcount} aliases")

# 5. Master Jalur
print("\n[5/14] Importing master_jalur...")
ws = wb['08_MASTER_JALUR']
rows = get_rows(ws)
for r in rows:
    if not r.get('Kode Jalur'): continue
    cur.execute("""
        INSERT INTO fcc.master_jalur (kode, nama, tujuan, peruntukan, site, status)
        VALUES (%s, %s, %s, %s, %s, %s) ON CONFLICT (kode) DO NOTHING
    """, (r.get('Kode Jalur'), r.get('Nama Jalur'), r.get('Tujuan'),
          r.get('Tujuan', 'TRANSFER'), r.get('Site', 'PPA-BIB'), r.get('Status', 'ACTIVE')))
conn.commit()
print(f"  → {cur.rowcount} jalur")

# 6. Master Main Tank
print("\n[6/14] Importing master_main_tank...")
ws = wb['09_MASTER_MAIN_TANK']
rows = get_rows(ws)
for r in rows:
    if not r.get('Kode Main Tank'): continue
    cur.execute("""
        INSERT INTO fcc.master_main_tank (kode, nama, kapasitas_l, status)
        VALUES (%s, %s, %s, %s) ON CONFLICT (kode) DO NOTHING
    """, (r.get('Kode Main Tank'), r.get('Nama Main Tank'),
          r.get('Kapasitas (L)', 0), r.get('Status', 'ACTIVE')))
conn.commit()
print(f"  → {cur.rowcount} tanks")

# 7. Master Fuel Truck
print("\n[7/14] Importing master_fuel_truck...")
ws = wb['10_MASTER_FUEL_TRUCK']
rows = get_rows(ws)
for r in rows:
    if not r.get('Kode Fuel Truck'): continue
    cur.execute("""
        INSERT INTO fcc.master_fuel_truck (kode, nama, tipe, kapasitas_l, status)
        VALUES (%s, %s, %s, %s, %s) ON CONFLICT (kode) DO NOTHING
    """, (r.get('Kode Fuel Truck'), r.get('Nama Fuel Truck'),
          r.get('Tipe'), r.get('Kapasitas (L)', 0), r.get('Status', 'ACTIVE')))
conn.commit()
print(f"  → {cur.rowcount} trucks")

# 8. FT Mandar Ocean
print("\n[8/14] Importing ft_mandar_ocean...")
ws = wb['07_FT_MANDAR_OCEAN']
rows = get_rows(ws)
for r in rows:
    if not r.get('ID FT'): continue
    cur.execute("""
        INSERT INTO fcc.ft_mandar_ocean (id_ft, no_lambung, no_polisi, kapasitas_l,
            t2_depan_cm, t2_belakang_cm, expired_komisioning, masa_berlaku)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s) ON CONFLICT (id_ft) DO NOTHING
    """, (r.get('ID FT'), r.get('No Lambung'), r.get('No Polisi'),
          r.get('Kapasitas (L)', 0), r.get('Tera T2 Depan (cm)'),
          r.get('Tera T2 Belakang (cm)'), r.get('Expired Komisioning'),
          r.get('Masa Berlaku')))
conn.commit()
print(f"  → {cur.rowcount} FT Mandar Ocean")

# 9. Sounding Table (84K rows)
print("\n[9/14] Importing sounding_table (84K rows)...")
ws = wb['12_TABEL_SOUNDING']
batch = []
count = 0
total = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]: continue
    aset = ser(row[0])
    # Row format: aset, dip_cm, volume_l, ...
    if len(row) >= 3 and row[1] is not None and row[2] is not None:
        try:
            dip = round(float(row[1]), 1)
            vol = round(float(row[2]), 3)
            batch.append((aset, dip, vol, 'ACTIVE'))
            count += 1
        except (ValueError, TypeError): continue

    if len(batch) >= 5000:
        psycopg2.extras.execute_values(cur,
            "INSERT INTO fcc.sounding_table (aset, dip_cm, volume_l, status) VALUES %s ON CONFLICT DO NOTHING",
            batch)
        total += count
        count = 0
        batch = []
        if total % 20000 == 0:
            conn.commit()
            print(f"  → {total:,} rows...")

if batch:
    psycopg2.extras.execute_values(cur,
        "INSERT INTO fcc.sounding_table (aset, dip_cm, volume_l, status) VALUES %s ON CONFLICT DO NOTHING",
        batch)
    total += count
print(f"  → {total:,} sounding points")

# 10. Shift Route Config
print("\n[10/14] Importing shift_route_config...")
ws = wb['13_SHIFT_ROUTE_CONFIG']
rows = get_rows(ws)
for r in rows:
    if not r.get('ID Konfigurasi'): continue
    cur.execute("""
        INSERT INTO fcc.shift_route_config
        (tanggal, shift, jalur, main_tank, fm_akhir_shift_sebelumnya, fm_aktual_awal, status)
        VALUES (%s, %s, %s, %s, %s, %s, %s) ON CONFLICT DO NOTHING
    """, (r.get('Tanggal'), r.get('Shift'), r.get('Jalur'), r.get('Main Tank'),
          r.get('FM Akhir Shift Sebelumnya', 0), r.get('FM Aktual Awal Shift', 0),
          'DRAFT'))
conn.commit()
print(f"  → {cur.rowcount} configs")

# 11. App Config
print("\n[11/14] Importing app_config...")
ws = wb['28_APP_CONFIG']
rows = get_rows(ws)
for r in rows:
    if not r.get('Parameter'): continue
    cur.execute("""
        INSERT INTO fcc.app_config (parameter, nilai, tipe, keterangan, rahasia)
        VALUES (%s, %s, %s, %s, %s) ON CONFLICT (parameter) DO UPDATE SET nilai=EXCLUDED.nilai
    """, (r.get('Parameter'), str(r.get('Nilai') or ''),
          r.get('Tipe', 'string'), r.get('Keterangan'),
          str(r.get('Rahasia', 'TIDAK')).upper() == 'YA'))
conn.commit()
print(f"  → {cur.rowcount} configs")

# 12. Ref Lookup (from seed or 02_REFERENSI)
print("\n[12/14] Importing ref_lookup...")
ws = wb['02_REFERENSI']
rows = get_rows(ws)
for r in rows:
    jenis = r.get('Jenis') or r.get('Lookup') or ''
    kode = r.get('Kode') or r.get('Nilai') or ''
    label = r.get('Label') or kode
    if not jenis or not kode: continue
    cur.execute("""
        INSERT INTO fcc.ref_lookup (jenis, kode, label, urutan, aktif)
        VALUES (%s, %s, %s, %s, true) ON CONFLICT DO NOTHING
    """, (jenis, kode, label, 0))
conn.commit()
print(f"  → {cur.rowcount} lookup entries")

conn.commit()

# Summary
print("\n" + "="*60)
print("IMPORT COMPLETE")
print("="*60)
for t in ['app_user','master_vendor','master_unit','unit_alias',
         'master_jalur','master_main_tank','master_fuel_truck','ft_mandar_ocean',
         'sounding_table','shift_route_config','app_config','ref_lookup']:
    cur.execute(f"SELECT count(*) FROM fcc.{t}")
    c = cur.fetchone()[0]
    print(f"  fcc.{t:25s} {c:>10,}")

conn.close()
print("\n✓ V6 migration complete")