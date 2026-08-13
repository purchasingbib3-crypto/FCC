from __future__ import annotations

import io
import re
import tempfile
from contextvars import ContextVar
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Any, Iterable

try:
    import xlrd  # legacy BIFF .xls reader; required in production requirements
except ImportError:  # keep .xlsx/backend startup healthy if optional dep is temporarily absent
    xlrd = None
from openpyxl import load_workbook

try:
    from python_calamine import CalamineWorkbook
except ImportError:
    CalamineWorkbook = None

_PARSER_ENGINE: ContextVar[str] = ContextVar("fcc_excel_parser_engine", default="UNKNOWN")

def current_parser_engine() -> str:
    return _PARSER_ENGINE.get()


SAP_MB51_MOVEMENT_TYPES = {"201", "202", "261", "262"}


@dataclass
class ParsedRow:
    source: str
    tanggal: date
    alias_unit: str
    # liter is retained as the exact source quantity for backward-compatible audit.
    # SAP MB51 keeps its original signed quantity here (GI negative, reversal positive).
    liter: float
    quantity_source_l: float = 0.0
    volume_net_l: float = 0.0
    shift: str = ""
    storage_location: str = ""
    source_row: int = 0
    source_format: str = ""
    source_record_id: str = ""
    movement_type: str = ""
    material: str = ""
    uom: str = ""


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ").strip()).upper()


def normalize_unit(value: Any) -> str:
    text = str(value or "").replace("\xa0", " ").upper().strip()
    text = re.sub(r"[\s\-_.\\/]+", "", text)
    return re.sub(r"[^A-Z0-9]", "", text)




def normalize_identifier(value: Any) -> str:
    """Preserve business identifiers as strings without Excel's trailing .0 artifact."""
    if value in (None, ""):
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).replace("\xa0", " ").strip()
    if re.fullmatch(r"-?\d+\.0+", text):
        return text.split(".", 1)[0]
    return text

def to_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "")
    neg = text.startswith("(") and text.endswith(")")
    if neg:
        text = text[1:-1]
    text = re.sub(r"[^0-9,.-]", "", text)
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return 0.0
    return -abs(number) if neg else number


def to_date(value: Any, datemode: int = 0) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            # Excel 1900 date system.  This is correct for the FCC files supplied.
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
        except Exception:
            if xlrd is not None:
                try:
                    return xlrd.xldate_as_datetime(value, datemode).date()
                except Exception:
                    return None
            return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def _find(headers: list[str], aliases: Iterable[str], required: bool = True) -> int | None:
    normalized = {normalize_header(h): i for i, h in enumerate(headers)}
    for alias in aliases:
        key = normalize_header(alias)
        if key in normalized:
            return normalized[key]
    if required:
        raise ValueError(f"Header tidak ditemukan: {', '.join(aliases)}")
    return None


def _read_matrix_calamine(content: bytes, filename: str) -> list[list[Any]]:
    if CalamineWorkbook is None:
        raise RuntimeError("python-calamine unavailable")
    suffix = ".xlsx" if filename.lower().endswith(".xlsx") else ".xls"
    temp_path = ""
    try:
        with tempfile.NamedTemporaryFile(prefix="fcc-import-", suffix=suffix, delete=False) as handle:
            handle.write(content)
            temp_path = handle.name
        workbook = CalamineWorkbook.from_path(temp_path)
        names = list(workbook.sheet_names)
        if not names:
            return []
        return workbook.get_sheet_by_name(names[0]).to_python(skip_empty_area=False)
    finally:
        if temp_path:
            try:
                import os
                os.unlink(temp_path)
            except OSError:
                pass


def read_matrix(content: bytes, filename: str) -> tuple[list[list[Any]], int]:
    """Read first sheet using the fastest available engine.

    V12.3 prefers python-calamine (Rust-backed) for both XLS/XLSX.  The existing
    xlrd/openpyxl readers remain deterministic fallbacks, so production stays
    functional even if the optional fast engine is temporarily unavailable.
    """
    lower = filename.lower()
    if CalamineWorkbook is not None:
        try:
            matrix = _read_matrix_calamine(content, filename)
            _PARSER_ENGINE.set("CALAMINE")
            return matrix, 0
        except Exception:
            # Fallback is intentional: parser speed must never reduce reliability.
            pass

    if lower.endswith(".xls") and not lower.endswith(".xlsx"):
        if xlrd is None:
            raise RuntimeError("File .xls membutuhkan python-calamine atau xlrd. Install dependency dari 06_env/requirements.txt lalu restart backend.")
        try:
            book = xlrd.open_workbook(file_contents=content, on_demand=True)
            sheet = book.sheet_by_index(0)
            matrix = [sheet.row_values(i) for i in range(sheet.nrows)]
            datemode = int(book.datemode)
            book.release_resources()
            _PARSER_ENGINE.set("XLRD")
            return matrix, datemode
        except Exception as exc:
            raise RuntimeError(f"Gagal membuka workbook XLS: {exc}") from exc

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        matrix = [list(row) for row in ws.iter_rows(values_only=True)]
        _PARSER_ENGINE.set("OPENPYXL")
        return matrix, 0
    except Exception as exc:
        raise RuntimeError(f"Gagal membuka workbook XLSX: {exc}") from exc


def _reconciliation_header_row(matrix: list[list[Any]], source: str) -> tuple[int, list[str], str]:
    """Find actual table header and detect the source format.

    Supported production contracts:
      * SS6_REFUELING: Transaction ID, Unit, Date, Shift, Vol, ...
      * SAP_MB51: Posting Date, Qty in Un. of Entry, Movement Type, Order/Text, ...
      * SAP_DIRECT: legacy summarized SAP export with a direct Unit column.
    """
    source = source.upper()
    for index, raw in enumerate(matrix[:50]):
        headers = [normalize_header(value) for value in raw]
        available = set(headers)
        if source == "SS6":
            if (
                any(x in available for x in {"DATE", "TANGGAL"})
                and any(x in available for x in {"UNIT", "NO LAMBUNG", "UNIT SS6"})
                and any(x in available for x in {"VOL", "VOLUME", "LITER"})
            ):
                return index, headers, "SS6_REFUELING"
        else:
            has_date = any(x in available for x in {"POSTING DATE", "TANGGAL", "DOCUMENT DATE"})
            has_qty = any(x in available for x in {"QTY", "QTY IN UN. OF ENTRY", "QUANTITY"})
            has_direct_unit = any(x in available for x in {"UNIT SAP FIX", "NO LAMBUNG SAP", "UNIT SAP", "UNIT"})
            has_mb51 = "MOVEMENT TYPE" in available and ("ORDER" in available or "TEXT" in available)
            if has_date and has_qty and has_mb51:
                return index, headers, "SAP_MB51"
            if has_date and has_qty and has_direct_unit:
                return index, headers, "SAP_DIRECT"
    raise ValueError(f"Header {source} tidak ditemukan pada 50 baris pertama workbook")


def _clean_sap_order_unit(value: Any) -> str:
    """Extract the equipment/unit code from SAP MB51 Order.

    Real FCC SAP examples include H78185-1201B, D85190-1201B, LT52506-1201,
    and variants with 120B/1201D/1201BT/trailing periods.  Only the known plant
    suffix at the end is removed; the actual unit code is never guessed.
    """
    text = str(value or "").strip().upper()
    if not text:
        return ""
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[-_.]?(?:1201BT|1201B|1201D|1201|120B)\.?$", "", text, flags=re.IGNORECASE)
    return text.strip(" -_.")


def _clean_sap_text_unit(value: Any) -> str:
    """Extract unit from SAP MB51 Text for 201/202 postings.

    Typical values are LV-5085.KM-55048 or BUS-1112.KM-216936.  If no KM/HM
    meter suffix exists, the full text is treated as the unit alias.
    """
    text = str(value or "").strip().upper()
    if not text:
        return ""
    match = re.search(r"[.]\s*(?:KM|HM)\s*[-_: ]*\d", text, flags=re.IGNORECASE)
    if match:
        text = text[: match.start()]
    return text.strip(" -_.")


def extract_sap_mb51_unit(movement_type: Any, order_value: Any, text_value: Any) -> str:
    movement = str(movement_type or "").strip()
    if movement in {"261", "262"}:
        return _clean_sap_order_unit(order_value)
    if movement in {"201", "202"}:
        return _clean_sap_text_unit(text_value)
    return ""


def _cell(raw: list[Any], index: int | None) -> Any:
    if index is None or index < 0 or index >= len(raw):
        return None
    return raw[index]


def parse_reconciliation_file(content: bytes, filename: str, source: str) -> list[ParsedRow]:
    source = source.upper().strip()
    if source not in {"SS6", "SAP"}:
        raise ValueError("source harus SS6 atau SAP")
    matrix, datemode = read_matrix(content, filename)
    if not matrix:
        return []
    header_index, headers, source_format = _reconciliation_header_row(matrix, source)

    if source == "SS6":
        date_col = _find(headers, ["DATE", "TANGGAL"])
        unit_col = _find(headers, ["UNIT", "NO LAMBUNG", "UNIT SS6"])
        qty_col = _find(headers, ["VOL", "VOLUME", "LITER"])
        shift_col = _find(headers, ["SHIFT"], False)
        sloc_col = _find(headers, ["GAS STATION", "STORAGE LOCATION", "SLOC", "STORAGE"], False)
        id_col = _find(headers, ["TRANSACTION ID", "ID", "NO VOUCHER"], False)
        material_col = _find(headers, ["MATERIAL"], False)
        uom_col = _find(headers, ["UOM", "UNIT OF ENTRY"], False)
        movement_col = order_col = text_col = doc_col = item_col = None
    elif source_format == "SAP_MB51":
        date_col = _find(headers, ["POSTING DATE", "TANGGAL", "DOCUMENT DATE"])
        qty_col = _find(headers, ["QTY IN UN. OF ENTRY", "QTY", "QUANTITY"])
        movement_col = _find(headers, ["MOVEMENT TYPE"])
        order_col = _find(headers, ["ORDER"], False)
        text_col = _find(headers, ["TEXT"], False)
        sloc_col = _find(headers, ["STORAGE LOCATION", "SLOC", "STORAGE"], False)
        doc_col = _find(headers, ["MATERIAL DOCUMENT"], False)
        item_col = _find(headers, ["MATERIAL DOC.ITEM", "MATERIAL DOC. ITEM", "MATERIAL DOCUMENT ITEM"], False)
        material_col = _find(headers, ["MATERIAL"], False)
        uom_col = _find(headers, ["UNIT OF ENTRY", "BASE UNIT OF MEASURE", "UOM"], False)
        unit_col = shift_col = id_col = None
    else:  # SAP_DIRECT backward compatibility
        date_col = _find(headers, ["POSTING DATE", "TANGGAL", "DOCUMENT DATE"])
        unit_col = _find(headers, ["UNIT SAP FIX", "NO LAMBUNG SAP", "UNIT SAP", "UNIT"])
        qty_col = _find(headers, ["QTY", "QTY IN UN. OF ENTRY", "QUANTITY"])
        shift_col = None
        sloc_col = _find(headers, ["STORAGE LOCATION", "SLOC", "STORAGE"], False)
        material_col = _find(headers, ["MATERIAL"], False)
        uom_col = _find(headers, ["UNIT OF ENTRY", "BASE UNIT OF MEASURE", "UOM"], False)
        doc_col = _find(headers, ["MATERIAL DOCUMENT"], False)
        item_col = _find(headers, ["MATERIAL DOC.ITEM", "MATERIAL DOC. ITEM"], False)
        movement_col = _find(headers, ["MOVEMENT TYPE"], False)
        order_col = text_col = id_col = None

    output: list[ParsedRow] = []
    for row_number, raw in enumerate(matrix[header_index + 1 :], start=header_index + 2):
        if not raw or all(v in (None, "") for v in raw):
            continue
        tanggal = to_date(_cell(raw, date_col), datemode)
        liter = to_number(_cell(raw, qty_col))
        if not tanggal or liter == 0:
            continue

        movement = str(_cell(raw, movement_col) or "").strip() if movement_col is not None else ""
        if source_format == "SAP_MB51":
            if movement not in SAP_MB51_MOVEMENT_TYPES:
                continue
            alias_raw = extract_sap_mb51_unit(movement, _cell(raw, order_col), _cell(raw, text_col))
        else:
            alias_raw = str(_cell(raw, unit_col) or "").strip()
        if not normalize_unit(alias_raw):
            continue

        uom = str(_cell(raw, uom_col) or "").strip().upper() if uom_col is not None else ""
        # Reconciliation is fuel-volume based.  If SAP/SS6 explicitly provides a
        # unit of measure, non-litre quantities are not silently mixed into fuel.
        if uom and uom not in {"L", "LTR", "LITER", "LITRE"}:
            continue

        if source == "SS6":
            source_record_id = normalize_identifier(_cell(raw, id_col)) if id_col is not None else ""
        else:
            doc = normalize_identifier(_cell(raw, doc_col)) if doc_col is not None else ""
            item = normalize_identifier(_cell(raw, item_col)) if item_col is not None else ""
            source_record_id = f"{doc}/{item}" if doc and item else (doc or "")

        # Canonical reconciliation direction:
        # - SS6 refueling volume is positive usage.
        # - SAP MB51 keeps signed source quantity for audit, while comparable net
        #   volume reverses the sign (261/201 issue negative -> positive usage;
        #   262/202 reversal positive -> negative usage).
        # - Legacy SAP_DIRECT historically relied on absolute usage values; keep
        #   that compatibility without changing MB51 semantics.
        volume_net_l = liter if source == "SS6" else (-liter if source_format == "SAP_MB51" else abs(liter))

        output.append(
            ParsedRow(
                source=source,
                tanggal=tanggal,
                alias_unit=alias_raw,
                liter=liter,
                quantity_source_l=liter,
                volume_net_l=volume_net_l,
                shift=str(_cell(raw, shift_col) or "").strip() if shift_col is not None else "",
                storage_location=str(_cell(raw, sloc_col) or "").strip() if sloc_col is not None else "",
                source_row=row_number,
                source_format=source_format,
                source_record_id=source_record_id,
                movement_type=movement,
                material=normalize_identifier(_cell(raw, material_col)) if material_col is not None else "",
                uom=uom,
            )
        )
    return output


def parse_ss6_export(content: bytes, filename: str) -> list[dict[str, Any]]:
    matrix, datemode = read_matrix(content, filename)
    if not matrix:
        return []
    header_index, headers, _ = _reconciliation_header_row(matrix, "SS6")
    idx = {h: i for i, h in enumerate(headers)}

    def col(*names: str, required: bool = False) -> int | None:
        for name in names:
            if normalize_header(name) in idx:
                return idx[normalize_header(name)]
        if required:
            raise ValueError(f"Header SS6 export tidak ditemukan: {names}")
        return None

    id_col = col("TRANSACTION ID", "ID", "NO VOUCHER")
    unit_col = col("UNIT", required=True)
    date_col = col("DATE", "TANGGAL", required=True)
    shift_col = col("SHIFT")
    time_col = col("TIME", "JAM")
    vol_col = col("VOL", "VOLUME", "LITER", required=True)
    hm_col = col("HM")
    gas_col = col("GAS STATION", "FUEL SOURCE")
    loc_col = col("LOCATION")
    fm_col = col("FM", "FUELMAN")
    input_col = col("INPUT BY")
    material_col = col("MATERIAL")

    output: list[dict[str, Any]] = []
    for row_index, raw in enumerate(matrix[header_index + 1 :], start=header_index + 2):
        if not raw or all(v in (None, "") for v in raw):
            continue
        tanggal = to_date(_cell(raw, date_col), datemode)
        unit_raw = str(_cell(raw, unit_col) or "").strip()
        volume = to_number(_cell(raw, vol_col))
        if not tanggal or not unit_raw or volume <= 0:
            continue
        txid = str(_cell(raw, id_col) or "").strip() if id_col is not None else ""
        shift_raw = str(_cell(raw, shift_col) or "").strip() if shift_col is not None else ""
        shift_number = "2" if "2" in shift_raw else "1"
        fingerprint = "|".join(
            [tanggal.isoformat(), shift_number, normalize_unit(unit_raw), f"{volume:.3f}", str(_cell(raw, time_col) or "")]
        )
        output.append(
            {
                "row_id": txid or fingerprint,
                "transaction_id": txid,
                "date": tanggal.isoformat(),
                "shift": f"SHIFT_{shift_number}",
                "unit_raw": unit_raw,
                "unit_normalized": normalize_unit(unit_raw),
                "volume_l": volume,
                "time": str(_cell(raw, time_col) or "").strip() if time_col is not None else "",
                "hm": to_number(_cell(raw, hm_col)) if hm_col is not None else 0,
                "gas_station": str(_cell(raw, gas_col) or "").strip() if gas_col is not None else "",
                "location": str(_cell(raw, loc_col) or "").strip() if loc_col is not None else "",
                "fuelman": str(_cell(raw, fm_col) or "").strip() if fm_col is not None else "",
                "input_by": str(_cell(raw, input_col) or "").strip() if input_col is not None else "",
                "material": str(_cell(raw, material_col) or "").strip() if material_col is not None else "",
                "source_row": row_index,
            }
        )
    return output
